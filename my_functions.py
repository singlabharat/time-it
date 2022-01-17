import json
import requests
import pandas
import openpyxl
import random
import datetime
import time

from requests.api import get

num = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]


def get_fact():
    response = requests.get(
        "https://uselessfacts.jsph.pl/today.json?language=en", verify=False)
    json_response = response.json()
    fact = json_response["text"]
    return fact


def set_quote(date, quote, author):
    path = "quote.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    date_cell = sheet_obj.cell(row=2, column=1)
    quote_cell = sheet_obj.cell(row=2, column=2)
    author_cell = sheet_obj.cell(row=2, column=3)
    date_cell.value = date
    quote_cell.value = quote
    author_cell.value = author
    wb_obj.save("quote.xlsx")


def get_qoute():
    current_time = datetime.datetime.now()
    current_date = str(current_time.day) + '-' + \
        str(current_time.month) + '-' + str(current_time.year)
    db = pandas.read_excel('quote.xlsx')
    my_array = db.to_numpy()
    quote_added = bool
    for i in range(len(my_array)):
        if my_array[i][0] == current_date:
            quote = my_array[i][1]
            author = my_array[i][2]
            return_list = [quote, author]
            quote_added = True
        else:
            quote_added = False
    if quote_added == False:
        response = requests.get("https://quotes.rest/qod?language=en", verify=False)
        json_response = response.json()
        quote = json_response["contents"]["quotes"][0]["quote"]
        author = json_response["contents"]["quotes"][0]["author"]
        set_quote(current_date, quote, author)
        return_list = [quote, author]
    return return_list


def set_yt_vid(date, link):
    path = "yt_vid.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    date_cell = sheet_obj.cell(row=2, column=1)
    link_cell = sheet_obj.cell(row=2, column=2)
    date_cell.value = date
    link_cell.value = link
    wb_obj.save("yt_vid.xlsx")


def get_youtube_vid():
    current_time = datetime.datetime.now()
    current_date = str(current_time.day) + '-' + str(current_time.month) + '-' + str(current_time.year)
    db = pandas.read_excel('yt_vid.xlsx')
    my_array = db.to_numpy()
    link = ""
    link_added = False
    for i in range(len(my_array)):
        if my_array[i][0] == current_date:
            link = my_array[i][1]
            link_added = True
        else:
            link_added = False
    if link_added == False:
        response = requests.get("https://www.googleapis.com/youtube/v3/search?q=stand+up+comedy&key=AIzaSyD0VsHXytSG_QarxOHFzO9MDYMJydLT-Ag&maxResults=10", verify=False)
        json_response = response.json()
        number = random.choice(num)
        vid_id = json_response["items"][number]['id']["videoId"]
        link = "https://www.youtube.com/embed/"+vid_id
        set_yt_vid(current_date, link)
    return link


def reminder_check(name):
    my_dict = dict()
    db = pandas.read_excel('radio_btn.xlsx')
    my_array = db.to_numpy()
    for i in range(len(my_array)):
        my_dict[my_array[i][0]] = my_array[i][1]
    answer = my_dict[name]
    return answer


def set_reminder(name, input_value):
    my_dict = {'water': 2, "meditate": 3,
               "exercise": 4, "rest": 5, "entertain": 6}
    path = "radio_btn.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    row = my_dict[name]
    obj_cell = sheet_obj.cell(row=row, column=2)
    obj_cell.value = input_value
    wb_obj.save("radio_btn.xlsx")


def check_time_table():
    my_dict_1 = dict()
    db = pandas.read_excel('time_table.xlsx')
    my_array = db.to_numpy()
    for i in range(len(my_array)):
        my_dict_1[my_array[i][0]] = my_array[i][1]
    return my_array



def set_time_table(slot, task, link):
    my_dict = {'9-10': 2, '10-11': 3, '11-12': 4, '12-13': 5, '13-14': 6, '14-15': 7,'15-16': 8, '16-17': 9, '17-18': 10, '18-19': 11, '19-20': 12, '20-21': 13}
    path = "time_table.xlsx"
    wb_obj = openpyxl.load_workbook(path.strip())
    sheet_obj = wb_obj.active
    row = my_dict[slot]
    task_cell = sheet_obj.cell(row=row, column=2)
    link_cell = sheet_obj.cell(row=row, column=3)
    task_cell.value = task
    link_cell.value = link
    wb_obj.save("time_table.xlsx")

def get_study_hack():
    my_list = ["20 minutes exercise before an studying can improve performance.", "Chewing gum is actually a brain booster. It helps in keeping us focused.", "Take a short break of 45-50 minutes as your focus and will be lost after this period.", "Meditate, it will help you concentrate when studying, it reduces pre-exam stress and improves mental & physical health.", "Every time you read through a page, eat a piece of candy to reward yourself.", "Study for 25 min with 5 min breaks between each study session. After 4 cycles take a longer 30 min break. This improves productivity."]
    db = pandas.read_excel('study_hack.xlsx')
    my_array = db.to_numpy()
    current_time = datetime.datetime.now()
    current_date = str(current_time.day) + '-' + str(current_time.month) + '-' + str(current_time.year)

    for i in range(len(my_array)):
        if my_array[i][0] == current_date:
            fact_added = True
            hack = my_array[i][1]
        else:
            fact_added = False
    if fact_added == False:
        hack = random.choice(my_list)
        path = "study_hack.xlsx"
        wb_obj = openpyxl.load_workbook(path.strip())
        sheet_obj = wb_obj.active
        date_cell = sheet_obj.cell(row=2, column=1)
        hack_cell = sheet_obj.cell(row=2, column=2)
        date_cell.value = current_date
        hack_cell.value = hack
        wb_obj.save("study_hack.xlsx")
    return hack

def get_time():
    current_time = datetime.datetime.now()
    current_time = str(current_time.hour) + '-' + str(current_time.minute)
    return current_time